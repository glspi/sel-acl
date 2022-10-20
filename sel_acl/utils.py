"""sel_acl.utils"""
import re
import sys
from ipaddress import ip_network
from typing import Dict, List

from ciscoconfparse import CiscoConfParse
from openpyxl.styles import PatternFill
from openpyxl.workbook.views import BookView
from openpyxl.workbook.workbook import Workbook

from sel_acl.objs import ACE, ACL, CustomWorksheet, MigrationData, NexusACL


def get_acl_from_file(filename: str, name: str):
    if name == "" or name is None:
        return None

    acl_parser = CiscoConfParse(filename)
    search_for = rf"ip access-list extended {name}"

    try:
        my_acl = ACL(name=name, acl=acl_parser.find_objects(search_for)[0])
    except IndexError:
        print(f"Error loading acl: '{name}' from {filename}, not found?")
        my_acl = None

    return my_acl


def get_nexus_acl_from_file(filename: str, name: str):
    if name == "" or name is None:
        return None

    acl_parser = CiscoConfParse(filename, syntax="nxos")
    search_for = rf"IP access list {name}"

    try:
        my_acl = NexusACL(name=name, acl=acl_parser.find_objects(search_for)[0])
    except IndexError:
        print(f"Error loading acl: '{name}' from {filename}, not found?")
        my_acl = None

    return my_acl


def get_addrgroups_from_file(filename: str) -> Dict:
    addr_group_parser = CiscoConfParse(filename)
    search_for = r"object-group ip address"

    addr_groups = {}
    group_names = addr_group_parser.find_objects(search_for)
    for group in group_names:
        match = re.match(r"object-group\sip\saddress\s(.+)", group.text)
        if match:
            name = match.group(1).strip()
            addr_groups[name] = []
            for host in group.children:
                match = re.match(r"\shost-info\s(.+)", host.text)
                if not match:
                    match = re.match(
                        r"\s(\d+\.\d+\.\d+\.\d+)\s(\d+\.\d+\.\d+\.\d+)(?:\s+)?",
                        host.text,
                    )
                    if match:
                        subnet, mask = match.groups()
                        network = str(ip_network(f"{subnet}/{mask}", strict=False))
                        addr_groups[name].append(network)
                else:
                    ip = match.group(1).strip() + "/32"
                    addr_groups[name].append(ip)

    return addr_groups


def get_nexus_addrgroups_from_file(filename: str) -> Dict:
    addr_group_parser = CiscoConfParse(filename)
    search_for = r"object-group ip address"

    addr_groups = {}
    group_names = addr_group_parser.find_objects(search_for)
    for group in group_names:
        match = re.match(r"object-group\sip\saddress\s(.+)", group.text)
        if match:
            name = match.group(1).strip()
            addr_groups[name] = []
            for host in group.children:
                match = re.match(r"\s+\d+\shost\s(.+)", host.text)
                # breakpoint()
                if not match:
                    match = re.match(
                        r"\s+\d+\s(\d+\.\d+\.\d+\.\d+\/\d+)(?:\s+)?",
                        host.text,
                    )
                    # breakpoint()
                    if match:
                        network = match.group(1).strip()
                        addr_groups[name].append(network)
                else:
                    ip = match.group(1).strip() + "/32"
                    addr_groups[name].append(ip)

    return addr_groups


def get_nexus_portgroups_from_file(filename: str) -> Dict:
    port_group_parser = CiscoConfParse(filename)
    search_for = r"object-group ip port"

    port_groups = {}
    group_names = port_group_parser.find_objects(search_for)
    for group in group_names:
        match = re.match(r"object-group\sip\sport\s(.+)", group.text)
        if match:
            name = match.group(1).strip()
            port_groups[name] = []
            for port in group.children:
                match = re.match(r"\s+\d+\seq\s(.+)", port.text)
                if not match:
                    match = re.match(
                        r"\s+\d+\srange\s(\d+\s\d+).+",
                        port.text,
                    )
                    if match:
                        # breakpoint()
                        ports = f" range {str(match.groups()[0])}"  # include 'range'
                        port_groups[name].append(ports)
                else:
                    ports = str(match.groups()[0])
                    port_groups[name].append(ports)

    return port_groups


def get_portgroups_from_file(filename: str) -> Dict:
    port_group_parser = CiscoConfParse(filename)
    search_for = r"object-group ip port"

    port_groups = {}
    group_names = port_group_parser.find_objects(search_for)
    for group in group_names:
        match = re.match(r"object-group\sip\sport\s(.+)", group.text)
        if match:
            name = match.group(1).strip()
            port_groups[name] = []
            for port in group.children:
                match = re.match(r"\seq\s(.+)", port.text)
                if not match:
                    match = re.match(
                        r"\srange\s(\d+\s\d+).+",
                        port.text,
                    )
                    if match:
                        ports = str(match.group())  # include 'range'
                        port_groups[name].append(ports)
                else:
                    ports = str(match.groups()[0])
                    port_groups[name].append(ports)

    return port_groups


def get_initial_data(excel_filename: str, vlan: int, obj_groups: str):
    # Get ACL/Name/Subnet/Etc from Vlan Number
    ws = CustomWorksheet(excel_filename=excel_filename)
    my_row = ws.find_row_from_vlan(vlan)
    mig_data = ws.get_migration_data_from_row(my_row)

    if not mig_data:
        print(
            f"We aren't migrating this vlan ({vlan}) or it has no subnet! (check the spreadsheet.)"
        )
        sys.exit()

    addr_groups = get_addrgroups_from_file(obj_groups)
    port_groups = get_portgroups_from_file(obj_groups)
    # return ws, mig_data, addr_groups, port_groups
    return {
        "ws": ws,
        "mig_data": mig_data,
        "addr_groups": addr_groups,
        "port_groups": port_groups,
    }


def remove_self(vlan_name: str, ew_mig_data: List[MigrationData]):
    ew_mig_data2 = []
    for mig_data in ew_mig_data:
        if mig_data.vlan_name != vlan_name:
            ew_mig_data2.append(mig_data)
    return ew_mig_data2


def addr_groups_to_nexus(names: List[str], addr_groups: Dict[str, Dict[str, List]]):
    groups = {}
    for name in names:
        if name in addr_groups:
            items = addr_groups[name]
            groups[name] = f"object-group ip address {name}\n"
            for item in items:
                if item.endswith("/32"):
                    groups[name] += f" host {item.replace('/32', '')}\n"
                else:
                    groups[name] += f" {item}\n"
        else:
            groups[name] = f"Could not find {name}!"
            print(
                f"Could not find address group: {name} but ACL refers to it...skipping."
            )

    return groups


def port_groups_to_nexus(names: List[str], port_groups: Dict[str, Dict[str, List]]):
    groups = {}
    for name in names:
        if name in port_groups:
            items = port_groups[name]
            groups[name] = f"object-group ip port {name}\n"
            for item in items:
                if item.startswith(" range"):
                    _ = item.replace(" range", "")
                    ports = _.strip()
                    port_start, port_end = ports.split()
                    groups[name] += f" range {port_start} {port_end}\n"
                else:
                    groups[name] += f" eq {item}\n"
        else:
            groups[name] = " Could not find!"
            print(
                f"Could not find address group: {name} but ACL refers to it...skipping."
            )

    return groups


def trim_acls(acls: List[ACL], mig_data, addr_groups, port_groups):
    acl_filename = input("Filename for existing NEXUS ACL's to compare with: ")

    for acl in acls:
        if acl.name.startswith("From"):
            direction = "From"  # IN
        elif acl.name.startswith("To"):
            direction = "To"  # OUT
        else:
            print(f"Could not determine ACL direction (to/from), ACL: {acl.name}")
            sys.exit()
        existing_acl_name = f"{direction}-{mig_data.tenant}-Tenant"
        acl_nxos = get_nexus_acl_from_file(
            filename=acl_filename, name=existing_acl_name
        )

        if acl_nxos:
            new_acl, removed_aces = trim_acl(
                acl_1=acl,
                acl_nxos=acl_nxos,
                addr_groups=addr_groups,
                port_groups=port_groups,
            )
            output_to_file(
                acl=new_acl,
                name=new_acl.name,
                addr_groups=addr_groups,
                port_groups=port_groups,
                removed=removed_aces,
            )
        else:
            print(f"{existing_acl_name} not found, nothing to compare to!")
            sys.exit()


def trim_acl(
    acl_1,
    acl_nxos,
    addr_groups,
    port_groups,
):

    obj_filename = "North-South-Object-Groups.ios"
    addr_groups_nxos = get_nexus_addrgroups_from_file(filename=obj_filename)
    port_groups_nxos = get_nexus_portgroups_from_file(filename=obj_filename)

    acl_1.set_cidrs_ports(addr_groups=addr_groups, port_groups=port_groups)
    acl_nxos.set_cidrs_ports(addr_groups=addr_groups_nxos, port_groups=port_groups_nxos)

    removed = []
    overlaps = acl_1.compare_with(acl_nxos)
    new_name = f"trimmed--{acl_1.name}"
    new_acl = ACL(name=new_name)
    for ace in acl_1.aces:
        remove = False
        for overlap in overlaps:
            if ace == overlap[0]:
                remove = True
                removed.append((ace.output_cidr(), overlap[1].output_cidr()))
        if not remove:
            new_acl.aces.append(ace)
    return new_acl, removed


def output_to_file(
    acl: ACL, name: str, addr_groups, port_groups, removed: List[str] = None
):
    output = acl.output_cidr(name=name)
    addr_group_names, port_group_names = acl.obj_groups()

    addr_groups = addr_groups_to_nexus(addr_group_names, addr_groups)
    port_groups = port_groups_to_nexus(port_group_names, port_groups)

    if removed:
        output += "\n\nRemoved ACE's: \n"
        for ace in removed:
            output += f"{ace[0]:<60} matched: {ace[1]}\n"

    output += "\n\nAddress-Groups: \n"
    for obj_group in addr_groups.values():
        output += obj_group

    output += "\n\nPort-Groups: \n"
    for group in port_groups.values():
        output += group

    filename = name + ".ios"
    with open(filename, "w") as fout:
        fout.write(output)
    print(f"File created at: {filename}\n")


def ns_ew_combined(ws, mig_data, acl, addr_groups, port_groups = None):

    ew_mig_data = ws.get_tenant_rows(tenant=mig_data.tenant)
    if ew_mig_data:
        ew_mig_data = remove_self(vlan_name=mig_data.vlan_name, ew_mig_data=ew_mig_data)

        if acl.name.startswith("From"):
            direction = "in"
        elif acl.name.startswith("To"):
            direction = "out"
        else:
            print("Could not determine ACL direction (to/from).")
            sys.exit()

        ew_aces, ew_contracts, ew_supernets = ew_checker(
            ew_mig_data=ew_mig_data,
            acl=acl,
            addr_groups=addr_groups,
            my_mig_data=mig_data,
            direction=direction,
            port_groups=port_groups
        )
        if ew_supernets:
            print(f"\n\nSUPERnet rules found in {acl.name}:")
            print(f"{'Overlapping VLAN':<60}:\t{' ACL: ACE Entry'}")
            print(
                "----------------------------------------------------------------------------------------"
            )
            for supernet in ew_supernets:
                vlan = f"{supernet['vlan_name']}  ({supernet['subnet']})"
                print(f"{vlan:<60}:\t{supernet['match']}")

        return ew_aces, ew_contracts


def create_contracts(ew_aces, ew_contracts, acl, filename):
    print(f"\n\nEast/West Contracts found in {acl.name}:")
    print("---------------------")

    if ew_contracts:
        for ace in ew_aces:
            if ace in acl.aces:
                print(ace.output_cidr())
        create_contract_file(contracts=ew_contracts, filename=filename)


def output_and_remark_acl(ew_aces, acl, name: str, addr_groups, port_groups):
    # Insert remarks so we can find these to delete after migrations
    for ace in ew_aces:
        if ace in acl.aces:
            index = acl.aces.index(ace)
            acl.aces.insert(
                index,
                ACE(
                    remark=f"### EAST/WEST BELOW, REMOVE ME LATER: {ace.output_cidr()}"
                ),
            )

    if ew_aces:
        print(f"\n\nEast/West Rules found in {acl.name}:")
        print("---------------------")
        for ace in ew_aces:
            if ace in acl.aces:
                print(ace.output_cidr())

    output_to_file(
        acl=acl,
        name=f"New-NS-w-EW-Remarked-{acl.name}",
        addr_groups=addr_groups,
        port_groups=port_groups,
    )


def run_cidr(acls: List[ACL], addr_groups, port_groups):
    print("CIDR Output..")
    print("--------------")
    for acl in acls:
        output_to_file(
            acl=acl,
            name=f"New-NS-{acl.name}",
            addr_groups=addr_groups,
            port_groups=port_groups,
        )


def run_ns(acls: List[ACL], addr_groups, port_groups, ws, mig_data):
    print("North/South...")
    print("--------------")

    for acl in acls:
        print(f"\nChecking {acl.name} and Tenant {mig_data.tenant} networks..\n")
        ew_aces, _ = ns_ew_combined(
            ws=ws, mig_data=mig_data, addr_groups=addr_groups, acl=acl,
        )
        output_and_remark_acl(
            acl=acl,
            name=f"New-NS-{acl.name}",
            addr_groups=addr_groups,
            port_groups=port_groups,
            ew_aces=ew_aces,
        )
    yesno = ""
    while yesno not in ("y", "n", "yes", "no"):
        yesno = input("Trim ACL? (y/n): ").lower()
    if yesno in ("y", "yes"):
        trim_acls(acls, mig_data, addr_groups, port_groups)


def run_contracts(acls: List[ACL], addr_groups, port_groups, ws, mig_data):
    print("East/West...")
    print("--------------")

    for acl in acls:
        print(f"\nChecking {acl.name}...")
        if acl.name.startswith("From"):
            direction = "in"
        elif acl.name.startswith("To"):
            direction = "out"
        else:
            print("Could not determine ACL direction (to/from).")
            sys.exit()
        ew_aces, ew_contracts = ns_ew_combined(
            ws=ws,
            mig_data=mig_data,
            addr_groups=addr_groups,
            acl=acl,
            port_groups=port_groups
        )
        create_contracts(
            ew_aces=ew_aces,
            ew_contracts=ew_contracts,
            acl=acl,
            filename=f"contracts-{mig_data.vlan_name}-{direction}.xlsx",
        )


def create_contract_file(filename: str, contracts: List[Dict[str, str]]):
    wb = Workbook()
    ws = wb.active
    headers = []
    for k, v in contracts[0].items():
        headers.append(k)
    ws.append(headers)
    start, end = ws.dimensions.split(":")
    headers = ws[start:end][0]
    for cell in headers:
        cell.fill = PatternFill("solid", fgColor="CCCCCC")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    view = [BookView(xWindow=0, yWindow=0, windowWidth=18140, windowHeight=15540)]
    wb.views = view
    # Get headers and put into dict so that sheet["headername"] = "column_letter"
    col_dict = {}
    for col in ws.iter_cols(1, ws.max_column):
        col_dict[col[0].value.upper()] = col[0].column_letter
    #
    # from rich.pretty import pprint
    # pprint(contracts)

    for contract in contracts:
        cell_dict = {}
        for header, value in contract.items():
            cell_dict[col_dict[header.upper()]] = value
        ws.append(cell_dict)

    wb.save(filename)
    print(f"\nContracts saved at '{filename}'.")


def ew_checker(
    ew_mig_data: List[MigrationData],
    acl: ACL,
    addr_groups: Dict[str, List[str]],
    my_mig_data: MigrationData,
    direction: str = "in",
    port_groups = None
):
    ew_aces = []
    ew_contracts = []
    ew_supernets = []
    for mig_data in ew_mig_data:
        for index, ace in enumerate(acl.aces):
            if ace.remark:
                continue
            if ace.tcp_flag == "established":
                continue
            for subnet in mig_data.subnet:
                if direction == "in":
                    src_dest_in = ace.destination_in(
                        ip_network(subnet, strict=False), addr_groups
                    )
                else:  # direction == "out"
                    src_dest_in = ace.source_in(
                        ip_network(subnet, strict=False), addr_groups
                    )

                if src_dest_in == "subnet":
                    ace.set_cidrs_ports(addr_groups=addr_groups, port_groups=port_groups)
                    contract = ace.to_contract(
                        acl=acl,
                        tenant=mig_data.tenant,
                        src_epg=my_mig_data.epg,
                        dst_epg=mig_data.epg,
                        src_application=my_mig_data.application,
                        dst_application=mig_data.application,
                        remark=acl.aces[index - 1].remark,
                    )
                    ew_aces.append(ace)
                    ew_contracts.append(contract)
                if src_dest_in == "supernet":
                    subnet_str = ""
                    for i, subnet_ in enumerate(mig_data.subnet):
                        if i == 0:
                            subnet_str += subnet_
                        else:
                            subnet_str += f", {subnet_}"
                    temp = {
                        "vlan_name": mig_data.vlan_name,
                        "subnet": subnet_str,
                        "match": f"{acl.name}: {ace.output_cidr(dns=False)}",
                    }
                    ew_supernets.append(temp)

    return ew_aces, ew_contracts, ew_supernets
