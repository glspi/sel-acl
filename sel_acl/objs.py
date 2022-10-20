"""sel_acl.objs"""
import ipaddress
import re
import socket
import sys
from dataclasses import dataclass, field
from ipaddress import ip_network
from itertools import product
from typing import List

import ciscoconfparse
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

ACL_RE_PATTERN = r"""
                ^\s+
                (?P<sequence>\d+\s)?                                                    # Sequence Number
                (
                (?P<remark>remark\s+.*)                                                 # Remark
                |
                (
                (?P<action>permit|deny)\s+                                              # Action
                (?P<protocol>\S+)\s+                                                    # Protocol
                (
                (?P<src_group>addrgroup\s\S+)                                           # Source addrgroup
                |(?P<src_wld>\d+\.\d+\.\d+\.\d+\s+\d+\.\d+\.\d+\.\d+)                   # OR Source Network
                |(?P<src_host>host\s+\d+\.\d+\.\d+\.\d+)                                # OR 'host x.x.x.x'
                |(?P<src_any>any)                                                       # OR 'any'
                )
                (\s+)?
                (
                    (?P<src_portgroup>portgroup\s\S+)                                   # OR Source portgroup
                    |   
                    (
                        (?P<src_port_match>(eq|neq|precedence|range|tos|lt|gt))\s+      # Source port Match ('eq' normally)
                        (
                        (?P<src_port_start>(?<=range\s)\S+)\s+(?P<src_port_end>\S+)     # Source port range 
                        |(?P<src_port>(?<!range\s)(?!\d+\.\d+\.\d+\.\d+).+)             # OR Source port (only)
                        )
                    )
                )?
                (\s+)   
                (
                (?P<dst_group>addrgroup\s\S+)                                           # Destination addrgroup
                |(?P<dst_wld>\d+\.\d+\.\d+\.\d+\s+\d+\.\d+\.\d+\.\d+)                   # OR Destination Network
                |(?P<dst_host>host\s+\d+\.\d+\.\d+\.\d+)                                # OR 'host x.x.x.x'
                |(?P<dst_any>any)                                                       # OR 'any'
                )
                (?:\s+)?
                (
                    (?P<dst_portgroup>portgroup\s\S+)                                   # OR Destination portgroup
                    |
                    (
                        (?P<dst_port_match>(eq|neq|precedence|range|tos|lt|gt))\s+      # Destination port Match ('eq' normally)
                        (
                        (?P<dst_port_start>(?<=range\s)\S+)\s+(?P<dst_port_end>\S+)     # Destination port range    
                        |(?P<dst_port>(?<!range\s).+)                                   # OR Destination port (only)
                        )
                    )
                )?
                (?:\s+)?
                (?P<flags_match>(match-any|match-all)\s+)?                              # match tcp flags
                (?P<tcp_flag>(((\+|-|)ack(\s*?)|(\+|-|)established(\s*?)|(\+|-|)fin(\s*?)|(\+|-|)fragments(\s*?)|(\+|-|)psh(\s*?)|(\+|-|)rst(\s*?)|(\+|-|)syn(\s*?)|urg(\s*?))+))?   # mostly just 'established'
                (?P<icmp_type>(administratively-prohibited|echo-reply|echo|mask-request|packet-too-big|parameter-problem|port-unreachable|redirect|router-advertisement|router-solicitation|time-exceeded|ttl-exceeded|unreachable))?    # icmp type
                (?P<log>(log-input|log))?                                               # log
                )
                )
                """

NXOS_RE_PATTERN = r"""
                ^\s+
                (?P<sequence>\d+\s)?                                                    # Sequence Number
                (
                (?P<remark>remark\s+.*)                                                 # Remark
                |
                (
                (?P<action>permit|deny)\s+                                              # Action
                (?P<protocol>\S+)\s+                                                    # Protocol
                (
                (?P<src_group>addrgroup\s\S+)                                           # Source addrgroup
                |(?P<src_cidr>\d+\.\d+\.\d+\.\d+\/\d+)                                  # OR Source Network
                |(?P<src_host>host\s+\d+\.\d+\.\d+\.\d+)                                # OR 'host x.x.x.x'
                |(?P<src_any>any)                                                       # OR 'any'
                )
                (\s+)?
                (
                    (?P<src_portgroup>portgroup\s\S+)                                   # OR Source portgroup
                    |   
                    (
                        (?P<src_port_match>(eq|neq|precedence|range|tos|lt|gt))\s+      # Source port Match ('eq' normally)
                        (
                        (?P<src_port_start>(?<=range\s)\S+)\s+(?P<src_port_end>\S+)     # Source port range 
                        |(?P<src_port>(?<!range\s)(?!\d+\.\d+\.\d+\.\d+).+)             # OR Source port (only)
                        )
                    )
                )?
                (\s+)   
                (
                (?P<dst_group>addrgroup\s\S+)                                           # Destination addrgroup
                |(?P<dst_cidr>\d+\.\d+\.\d+\.\d+/\d+)                                   # OR Destination Network
                |(?P<dst_host>host\s+\d+\.\d+\.\d+\.\d+)                                # OR 'host x.x.x.x'
                |(?P<dst_any>any)                                                       # OR 'any'
                )
                (?:\s+)?
                (
                    (?P<dst_portgroup>portgroup\s\S+)                                   # OR Destination portgroup
                    |
                    (
                        (?P<dst_port_match>(eq|neq|precedence|range|tos|lt|gt))\s+      # Destination port Match ('eq' normally)
                        (
                        (?P<dst_port_start>(?<=range\s)\S+)\s+(?P<dst_port_end>\S+)     # Destination port range    
                        |(?P<dst_port>(?<!range\s).+)                                   # OR Destination port (only)
                        )
                    )
                )?
                (?:\s+)?
                (?P<flags_match>(match-any|match-all)\s+)?                              # match tcp flags
                (?P<tcp_flag>(((\+|-|)ack(\s*?)|(\+|-|)established(\s*?)|(\+|-|)fin(\s*?)|(\+|-|)fragments(\s*?)|(\+|-|)psh(\s*?)|(\+|-|)rst(\s*?)|(\+|-|)syn(\s*?)|urg(\s*?))+))?   # mostly just 'established'
                (?P<icmp_type>(administratively-prohibited|echo-reply|echo|mask-request|packet-too-big|parameter-problem|port-unreachable|redirect|router-advertisement|router-solicitation|time-exceeded|ttl-exceeded|unreachable))?    # icmp type
                (?P<log>(log-input|log))?                                               # log
                )
                )
                """


def load_excel(filename: str) -> openpyxl.workbook.workbook.Workbook:
    try:
        tmp = openpyxl.load_workbook(filename)  # , read_only=True)
    except InvalidFileException as error:
        print(error)
        sys.exit()

    return tmp


class CustomWorksheet:
    def __init__(self, excel_filename):
        # Load excel and get first sheet/tab
        _ = load_excel(excel_filename)
        self.worksheet: Workbook = _.active

        # Get headers and put into dict so that sheet["headername"] = "column_letter"
        self.col_dict = {}
        for col in self.worksheet.iter_cols(1, self.worksheet.max_column):
            self.col_dict[col[0].value.upper()] = col[0].column_letter

    def get_migration_data_from_row(self, row):
        # Check if we even should migrate or return anything
        if self.worksheet[self.col_dict["MIGRATE TO ACI?"] + str(row)].value == "N":
            return None
        subnets = []
        subnet_val = self.worksheet[self.col_dict["IP-ADDRESS"] + str(row)].value
        if subnet_val:
            subnet_val = subnet_val.strip()
            subnet_secondary = subnet_val.split(",")
            for subnet in subnet_secondary:
                subnet = subnet.strip()
                subnet = subnet.replace(" secondary", "")
                try:
                    _ = ip_network(subnet, strict=False)
                    subnets.append(subnet)
                except ValueError as e:
                    print(
                        f"Error: Bad subnet in: {self.worksheet[self.col_dict['NAME'] + str(row)].value}",
                        end="",
                    )
                    print(f"\t\t{e}")
                    return None
        else:
            return None

        acl = self.worksheet[self.col_dict["ACCESS-LISTS"] + str(row)].value

        if acl is None:
            in_ = ""
            out = ""
        else:
            in_out = self.col_dict.get("ACCESS-LIST-IN")
            if in_out:
                in_ = self.worksheet[in_out + str(row)].value
                out = self.worksheet[self.col_dict["ACCESS-LIST-OUT"] + str(row)].value
            else:
                acl = acl.strip()
                try:
                    in_, out = acl.split(",")
                except ValueError:
                    try:
                        in_, out = acl.split(" ")
                    except ValueError as e:
                        in_ = ""
                        out = None
                        if acl.startswith("From"):
                            in_ = acl
                        elif acl.startswith("To"):
                            out = acl
                        else:
                            print(f"Your Access-Lists column in broken. ", end="")
                            print(f"Find and fix {acl}")
                            print(e)
                            sys.exit()

        data = {
            "vlan_name": self.worksheet[self.col_dict["NAME"] + str(row)].value,
            "acl_name_in": in_,
            "acl_name_out": out,
            "tenant": self.worksheet[self.col_dict["TENANT"] + str(row)].value,
            "subnet": subnets,
            "epg": self.worksheet[self.col_dict["EPG"] + str(row)].value,
            "application": self.worksheet[
                self.col_dict["APPLICATION"] + str(row)
            ].value,
        }

        for k, v in data.items():
            if v is None:
                continue
            else:
                if isinstance(v, str):
                    data[k] = v.strip()

        return MigrationData(**data)

    def find_row_from_vlan(self, vlan: int):
        # Get Row for vlan to be migrated
        my_row = False
        for cell in self.worksheet[self.col_dict["VLAN"]]:
            if cell.value == vlan:
                my_row = cell.row
        if not my_row:
            print("Invalid vlan? Not found..")
            sys.exit()

        return my_row

    def get_rows_from_column(self, column: str, exclude: str = None):
        rows = []
        for i, row in enumerate(self.worksheet[self.col_dict[column]]):
            if i == 0:
                continue
            value = row.value.strip() if row.value else None
            if not value:
                continue
            if value == exclude:
                continue
            if value not in rows:
                rows.append(value)
        return rows

    def get_tenant_rows(self, tenant: str):
        rows = []
        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.column_letter == self.col_dict["TENANT"]:
                    if cell.value == tenant:
                        rows.append(row)
        mig_data_list = []
        for row in rows:
            my_row = row[0].row  # row number
            my_vlan = self.get_migration_data_from_row(row=my_row)
            if my_vlan:
                mig_data_list.append(my_vlan)

        return mig_data_list

    # def get_subnets(self, exclude_subnet: str):
    #     subnets = []
    #     for i, row in enumerate(self.worksheet[self.col_dict["IP-ADDRESS"]]):
    #         if i == 0:
    #             continue
    #         subnet = row.value.strip() if row.value else None
    #         if subnet == exclude_subnet or None:
    #             continue
    #         subnets.append(subnet)
    #     return subnets


@dataclass()
class MigrationData:
    vlan_name: str
    acl_name_in: str
    acl_name_out: str
    tenant: str
    subnet: str
    epg: str
    application: str


@dataclass()
class ACE:
    sequence: str = None
    remark: str = None
    action: str = None
    protocol: str = None
    src_group: str = None
    src_wld: str = None
    src_host: str = None
    src_any: str = None
    src_portgroup: str = None
    src_port_match: str = None
    src_port_start: str = None
    src_port_end: str = None
    src_port: str = None
    dst_group: str = None
    dst_wld: str = None
    dst_host: str = None
    dst_any: str = None
    dst_portgroup: str = None
    dst_port_match: str = None
    dst_port_start: str = None
    dst_port_end: str = None
    dst_port: str = None
    flags_match: str = None
    tcp_flag: str = None
    icmp_type: str = None
    log: str = None

    # Custom/Created Variables
    src_cidr: str = None  # field(default_factory=str)
    dst_cidr: str = None  # field(default_factory=str)
    src_cidrs: List = field(default_factory=list)
    src_ports: List = field(default_factory=list)
    dst_cidrs: List = field(default_factory=list)
    dst_ports: List = field(default_factory=list)

    def __post_init__(self):
        # Initialize ACE in a more reusable way
        if self.sequence:
            self.sequence = self.sequence.strip()
        if self.remark:
            self.remark = self.remark.replace("remark ", "")
        if self.src_group:
            self.src_group = self.src_group.replace("addrgroup ", "")
        if self.src_host:
            self.src_host = self.src_host.replace("host ", "")
            self.src_host += "/32"
        if self.src_portgroup:
            self.src_portgroup = self.src_portgroup.replace("portgroup ", "")
        if self.dst_group:
            self.dst_group = self.dst_group.replace("addrgroup ", "")
        if self.dst_host:
            self.dst_host = self.dst_host.replace("host ", "")
            self.dst_host += "/32"
        if self.dst_portgroup:
            self.dst_portgroup = self.dst_portgroup.replace("portgroup ", "")

        # Create CIDR if needed
        try:
            if not self.src_any and self.src_wld and not self.src_cidr:
                src_net, src_wildcard = self.src_wld.split()
                self.src_cidr = (
                    src_net
                    + "/"
                    + str(
                        ipaddress.ip_network(
                            src_net + "/" + src_wildcard, strict=False
                        ).prefixlen
                    )
                )
            if not self.dst_any and self.dst_wld and not self.dst_cidr:
                dst_net, dst_wildcard = self.dst_wld.split()
                self.dst_cidr = (
                    dst_net
                    + "/"
                    + str(
                        ipaddress.ip_network(
                            dst_net + "/" + dst_wildcard, strict=False
                        ).prefixlen
                    )
                )
        except ValueError as e:
            # Now that strict=False this try/except block not needed...?
            if "has host bits set" in str(e):
                print(
                    "ACL is 'incorrect', host bits are set in network, CIDR could not be created."
                )
                print(self)

    def output_cidr(self, dns=True):
        output = " "
        fqdn_remark = ""

        if self.sequence:
            output += f"{self.sequence} "

        if self.remark:
            output += "remark " + self.remark
            return output

        output += f"{self.action} {self.protocol} "
        # SOURCE ADDRESS
        if self.src_group:
            output += f"addrgroup {self.src_group} "
        elif self.src_any:
            output += "any "
        elif self.src_host:
            output += f"{self.src_host} "
            if dns:
                fqdn_remark += " remark ### From: "
                try:
                    ip = self.src_host.replace("/32", "")
                    fqdn = socket.gethostbyaddr(ip)[0]
                    fqdn_remark += fqdn
                except socket.herror:
                    fqdn_remark += "UNKNOWN"

        elif self.src_cidr:
            output += f"{self.src_cidr} "
        else:
            pass
            # print(f"Error, no source found in ACE {self}.")
        # SOURCE PORT
        if self.src_portgroup:
            output += f"portgroup {self.src_portgroup} "
        elif self.src_port_match:
            output += f"{self.src_port_match} "
        if self.src_port_start and self.src_port_end:
            output += f"{self.src_port_start} {self.src_port_end} "
        elif self.src_port:
            output += f"{self.src_port} "

        # DESTINATION
        if self.dst_group:
            output += f"addrgroup {self.dst_group} "
        elif self.dst_any:
            output += "any "
        elif self.dst_host:
            output += f"{self.dst_host} "
            if dns:
                if fqdn_remark:
                    fqdn_remark += "\tTo: "
                else:
                    fqdn_remark += " remark ### To: "
                try:
                    ip = self.dst_host.replace("/32", "")
                    fqdn = socket.gethostbyaddr(ip)[0]
                    fqdn_remark += fqdn
                except socket.herror:
                    fqdn_remark += "UNKNOWN"

        elif self.dst_cidr:
            output += f"{self.dst_cidr} "
        else:
            pass
            # print(f"Error, no destination found in ACE {self.__dict__}.")
        # DESTINATION PORT
        if self.dst_portgroup:
            output += f"portgroup {self.dst_portgroup} "
        elif self.dst_port_match:
            output += f"{self.dst_port_match} "
        if self.dst_port_start and self.dst_port_end:
            output += f"{self.dst_port_start} {self.dst_port_end} "
        elif self.dst_port:
            output += f"{self.dst_port} "

        # EXTRAS (flags/icmp/log)
        if self.flags_match:
            output += f"{self.flags_match} "
        if self.tcp_flag:
            output += f"{self.tcp_flag} "
        if self.icmp_type:
            output += f"{self.icmp_type} "
        if self.log:
            output += f"{self.log} "

        if fqdn_remark:
            output = fqdn_remark + "\n" + output
        return output

    def load_addr_cidrs(self, addr_groups, group):
        cidrs = []
        my_cidrs = addr_groups.get(group)
        if my_cidrs:
            for network in my_cidrs:
                try:
                    _ = ip_network(network, strict=False)  # Make sure it's a cidr
                    cidrs.append(network)
                except ValueError as e:
                    print(f"ERR: {e}")
                    print("ACE is incorrect")
                    print(network)
                    print(self.output_cidr())
                    sys.exit()

        else:
            print(f"Error getting address group: {group}, ignoring...")

        return cidrs

    def load_ports(self, port_groups, group):
        ports = []
        my_ports = port_groups.get(group)
        if my_ports:
            for port in my_ports:
                ports.append(port)
        else:
            print(f"Error getting port group: {group}, ignoring...")

        return ports

    def set_cidrs_ports(self, addr_groups, port_groups):
        if not self.src_cidrs:
            self.src_cidrs = []
        if not self.src_ports:
            self.src_ports = []
        if not self.dst_cidrs:
            self.dst_cidrs = []
        if not self.dst_ports:
            self.dst_ports = []

        if self.remark:
            return

        # SOURCE ADDRESS
        if self.src_host:
            self.src_cidrs.append(self.src_host)
        elif self.src_any:
            self.src_cidrs.append("0.0.0.0/0")
        elif self.src_cidr:
            self.src_cidrs.append(self.src_cidr)
        elif self.src_group:
            self.src_cidrs += self.load_addr_cidrs(
                addr_groups=addr_groups, group=self.src_group
            )
        # SOURCE PORT
        if self.src_portgroup:
            self.src_ports += self.load_ports(
                port_groups=port_groups, group=self.src_portgroup
            )
        elif self.src_port_match:
            if self.src_port_match == "eq":
                self.src_ports.append(self.src_port)
            elif self.src_port_match == "range":
                try:
                    for port in range(
                        int(self.src_port_start), int(self.src_port_end) + 1
                    ):
                        self.src_ports.append(str(port))
                except ValueError as e:
                    self.src_ports.append(self.src_port_start)
                    self.src_ports.append(self.src_port_end)
            elif self.src_port_match == "gt":
                for port in range(int(self.src_port), 65535):
                    self.src_ports.append(port)
            elif self.src_port_match == "lt":
                for port in range(0, int(self.src_port) + 1):
                    self.src_ports.append(port)
            elif self.src_port_match == "neq":
                print("Don't use `neq`, slowing us down:")
                for port in range(0, 65535):
                    if int(self.src_port) != port:
                        self.src_ports.append(str(port))
            else:
                match = self.src_port_match
                self.src_ports.append(f"{match} {self.src_port}")

        # DESTINATION
        if self.dst_host:
            self.dst_cidrs.append(self.dst_host)
        elif self.dst_any:
            self.dst_cidrs.append("0.0.0.0/0")
        elif self.dst_cidr:
            self.dst_cidrs.append(self.dst_cidr)
        elif self.dst_group:
            self.dst_cidrs += self.load_addr_cidrs(
                addr_groups=addr_groups, group=self.dst_group
            )
        # DESTINATION PORT
        if self.dst_portgroup:
            self.dst_ports += self.load_ports(
                port_groups=port_groups, group=self.dst_portgroup
            )
        elif self.dst_port_match:
            if self.dst_port_match == "eq":
                self.dst_ports.append(self.dst_port)
            elif self.dst_port_match == "range":
                try:
                    for port in range(
                        int(self.dst_port_start), int(self.dst_port_end) + 1
                    ):
                        self.dst_ports.append(str(port))
                except ValueError as e:
                    self.dst_ports.append(self.dst_port_start)
                    self.dst_ports.append(self.dst_port_end)
            elif self.dst_port_match == "gt":
                for port in range(int(self.dst_port), 65535):
                    self.dst_ports.append(port)
            elif self.dst_port_match == "lt":
                for port in range(0, int(self.dst_port) + 1):
                    self.dst_ports.append(str(port))
            elif self.dst_port_match == "neq":
                print("Don't use `neq`, slowing us down:")
                for port in range(0, 65535):
                    if self.dst_port != str(port):
                        self.dst_ports.append(str(port))
            else:
                match = self.dst_port_match
                self.dst_ports.append(f"{match} {self.dst_port}")

    def addr_groups(self):
        names = []
        if self.src_group:
            names.append(self.src_group)
        if self.dst_group:
            names.append(self.dst_group)

        return names

    def port_groups(self):
        names = []
        if self.src_portgroup:
            names.append(self.src_portgroup)
        if self.dst_portgroup:
            names.append(self.dst_portgroup)

        return names

    def destination_in(self, subnet, addr_groups):
        if self.dst_group:
            my_destinations = addr_groups.get(self.dst_group)
            if my_destinations:
                for network in my_destinations:
                    try:
                        my_destination = ip_network(network, strict=False)
                    except ValueError as e:
                        return f"ERR: {e}"
                    if my_destination.subnet_of(subnet):
                        return "subnet"
                    if my_destination.supernet_of(subnet):
                        return "supernet"
            else:
                print(f"Error getting address group: {self.dst_group}, skipping...")
                return False
        else:
            if self.dst_host:
                my_destination = self.dst_host
            if self.dst_any:
                my_destination = "0.0.0.0/0"
            if self.dst_cidr:
                my_destination = self.dst_cidr
            try:
                my_destination = ip_network(my_destination, strict=False)
                if my_destination.subnet_of(subnet):
                    return "subnet"
                if my_destination.supernet_of(subnet):
                    return "supernet"
            except ValueError as e:
                return f"ERR: {e}"
            except UnboundLocalError as e:
                # my_destination was never found/assigned
                print(f"ACE is incorrect, cannot process: {self.__dict__}")

    def source_in(self, subnet, addr_groups):
        if self.src_group:
            my_sources = addr_groups.get(self.src_group)
            if my_sources:
                for network in my_sources:
                    try:
                        my_source = ip_network(network, strict=False)
                    except ValueError as e:
                        return f"ERR: {e}"
                    if my_source.subnet_of(subnet):
                        return "subnet"
                    if my_source.supernet_of(subnet):
                        return "supernet"
            else:
                print(f"Error getting address group: {self.src_group}, skipping...")
                return False
        else:
            if self.src_host:
                my_source = self.src_host
            if self.src_any:
                my_source = "0.0.0.0/0"
            if self.src_cidr:
                my_source = self.src_cidr
            try:
                my_source = ip_network(my_source, strict=False)
                if my_source.subnet_of(subnet):
                    return "subnet"
                if my_source.supernet_of(subnet):
                    return "supernet"
            except ValueError as e:
                return f"ERR: {e}"
            except UnboundLocalError as e:
                # my_destination was never found/assigned
                print(f"ACE is incorrect, cannot process: {self.__dict__}")

    def compare_cidrs(self, ace_2, attr):
        iters = []
        for cidr in getattr(self, attr):
            match = False
            for cidr_2 in getattr(ace_2, attr):
                net_1 = ip_network(cidr)
                net_2 = ip_network(cidr_2)
                if net_1.subnet_of(net_2):
                    match = True
            iters.append(match)

        if iters:
            if all(iters):
                return iters

    def compare_ports(self, ace_2, attr):
        iters = []
        ports_1 = getattr(self, attr)
        ports_2 = getattr(ace_2, attr)

        # 'same' if no ports
        if not ports_1:
            if not ports_2:
                return [True]

        # Loop through our ports
        for port_1 in ports_1:
            match = False
            if port_1 in ports_2:
                match = True
            iters.append(match)
        if iters:
            if all(iters):
                return iters

    def compare_with(self, ace_2):
        if self.remark or ace_2.remark:
            if self.remark == ace_2.remark:
                return True

        elif self.action == ace_2.action:
            if self.protocol == ace_2.protocol:
                matches = self.compare_cidrs(ace_2, "src_cidrs")
                if matches:
                    matches = self.compare_cidrs(ace_2, "dst_cidrs")
                    if matches:
                        matches = self.compare_ports(ace_2, "src_ports")
                        if matches:
                            matches = self.compare_ports(ace_2, "dst_ports")
                            if matches:
                                more_matches = [
                                    self.flags_match == ace_2.flags_match,
                                    self.tcp_flag == ace_2.tcp_flag,
                                    self.icmp_type == ace_2.icmp_type,
                                    self.log == ace_2.log,
                                ]
                                if all(more_matches):
                                    return True

    def to_contract(
        self, acl, tenant, src_epg, dst_epg, src_application, dst_application, remark
    ):
        source, destination, source_port, destination_port = "", "", "", ""

        # SOURCE
        if self.src_group:
            source = self.src_group
        elif self.src_host:
            source = self.src_host
        elif self.src_any:
            source = "0.0.0.0/0"
        elif self.src_cidr:
            source = self.src_cidr

        if self.src_portgroup:
            # source_port = self.src_portgroup
            source_port = ""
            for port in self.src_ports:
                if "range" in port:
                    _, _, start, end = port.split(" ")
                    source_port += f"{start} - {end},"
                else:
                    source_port += f"{port},"
            source_port = source_port.rstrip(",")
        # elif self.src_port_match:
        #     source_port = f"{self.src_port_match} "
        if self.src_port_start and self.src_port_end:
            source_port = f"{self.src_port_start} - {self.src_port_end}"
        elif self.src_port:
            source_port = f"{self.src_port}"

        # DESTINATION
        if self.dst_group:
            destination = self.dst_group
        elif self.dst_host:
            destination = self.dst_host
        elif self.dst_any:
            destination = "0.0.0.0/0"
        elif self.dst_cidr:
            destination = self.dst_cidr

        if self.dst_portgroup:
            #destination_port = self.dst_portgroup
            source_port = ""
            for port in self.dst_ports:
                if "range" in port:
                    _, _, start, end = port.split(" ")
                    destination_port += f"{start} - {end},"
                else:
                    destination_port += f"{port},"
            destination_port = destination_port.rstrip(",")
        # elif self.dst_port_match:
        #     destination_port_match = f"{self.dst_port_match} "
        elif self.dst_port_start and self.dst_port_end:
            destination_port = f"{self.dst_port_start} - {self.dst_port_end}"
        elif self.dst_port:
            destination_port = f"{self.dst_port}"

        if not remark:
            remark = ""

        contract = {
            "acl_name": acl.name,
            "action": self.action,
            "protocol": self.protocol,
            "src_aci": "",
            "src_tenant": tenant,
            "src_epg": src_epg,
            "src_application": src_application,
            "src_address": source,
            "src_port": source_port,
            "src_portgroup": self.src_portgroup,
            "dst_tenant": tenant,
            "dst_epg": dst_epg,
            "dst_application": dst_application,
            "dst_aci": "",
            "dst_address": destination,
            "dst_port": destination_port,
            "dst_portgroup": self.dst_portgroup,
            "flags": self.tcp_flag,
            "remark": remark,
            "contract": "",
            "access-list": "",
            "ace": "",
            "Notes": "",
            "Completed": "",
        }
        return contract


@dataclass()
class ACL:
    name: str
    acl: ciscoconfparse.models_cisco.IOSCfgLine = None
    aces: list[ACE] = field(default_factory=list)
    rec = re.compile(ACL_RE_PATTERN, re.X)
    # groups = None

    def __post_init__(self):
        if self.acl:
            for child in self.acl.children:
                results = self.rec.search(child.text)
                if not results:
                    print(child.text)
                    print("^^ERROR WITH ABOVE LINE:^^")
                else:
                    ace = ACE(**results.groupdict())

                    # Some rules might be 'eq port1 port2 port3' which should be split into extra rules
                    if (
                        ace.src_port_match == "eq"
                        or ace.dst_port_match == "eq"
                        or ace.src_port_match == "neq"
                        or ace.dst_port_match == "neq"
                    ):
                        if ace.src_port:
                            src_ports = ace.src_port.split()
                        else:
                            src_ports = None
                        if ace.dst_port:
                            dst_ports = ace.dst_port.split()
                        else:
                            dst_ports = None

                        if src_ports and not dst_ports:
                            for port in src_ports:
                                temp = results.groupdict()
                                temp["src_port"] = port
                                new_ace = ACE(**temp)
                                self.aces.append(new_ace)
                        elif not src_ports and dst_ports:
                            for port in dst_ports:
                                temp = results.groupdict()
                                temp["dst_port"] = port
                                new_ace = ACE(**temp)
                                self.aces.append(new_ace)
                        elif src_ports and dst_ports:
                            for rule in product(src_ports, dst_ports):
                                temp = results.groupdict()
                                temp["src_port"] = rule[0]
                                temp["dst_port"] = rule[1]
                                new_ace = ACE(**temp)
                                self.aces.append(new_ace)
                    else:
                        self.aces.append(ace)

    def __contains__(self, find_ace):
        for ace in self.aces:
            if ace == find_ace:
                return True

    def obj_groups(self):
        addr_names = self.addr_groups()
        portgroup_names = self.port_groups()
        return addr_names, portgroup_names

    def addr_groups(self):
        addr_names = []
        for ace in self.aces:
            names = ace.addr_groups()
            for name in names:
                if name is not None and name not in addr_names:
                    addr_names.append(name)
        return addr_names

    def port_groups(self):
        portgroup_names = []
        for ace in self.aces:
            names = ace.port_groups()
            for name in names:
                if name is not None and name not in portgroup_names:
                    portgroup_names.append(name)
        return portgroup_names

    def set_cidrs_ports(self, addr_groups, port_groups):
        for ace in self.aces:
            ace.set_cidrs_ports(addr_groups=addr_groups, port_groups=port_groups)

    def output_groups(self):
        addr_group_names = self.addr_groups()
        return addr_group_names

    def compare_with(self, acl_2):
        overlaps = []
        for ace in self.aces:
            for ace_2 in acl_2.aces:
                if ace.compare_with(ace_2):
                    overlaps.append((ace, ace_2))
                else:
                    pass  # Do nothing so far
        return overlaps

    def output_cidr(self, name: str):
        output = f"ip access-list {name}\n"
        for ace in self.aces:
            if ace.tcp_flag == "established":
                continue
            output += ace.output_cidr() + "\n"
        return output


@dataclass()
class NexusACL(ACL):
    rec = re.compile(NXOS_RE_PATTERN, re.X)

    def __post_init__(self):
        if self.acl:
            for child in self.acl.children:
                results = self.rec.search(child.text)
                if not results:
                    print(child.text)
                    print("^^ERROR WITH ABOVE LINE:^^")
                else:
                    ace = ACE(**results.groupdict())
                    self.aces.append(ace)
