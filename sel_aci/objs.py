import sys
from dataclasses import dataclass, field
from typing import List

import openpyxl
from jinja2 import Environment, PackageLoader, StrictUndefined
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

j2_env = None


# Jinja Environment
def set_j2_env(version):
    if version == "5.2":
        folder = "templates-5.2"
    else:
        folder = "templates"
    global j2_env
    j2_env = Environment(
        # loader=FileSystemLoader("sel_aci/templates"),
        loader=PackageLoader("sel_aci", folder),
        lstrip_blocks=True,
        trim_blocks=True,
        undefined=StrictUndefined,
    )


def load_excel(filename: str) -> openpyxl.workbook.workbook.Workbook:
    try:
        tmp = openpyxl.load_workbook(filename)  # , read_only=True)
    except InvalidFileException as error:
        print(error)
        sys.exit()

    return tmp


@dataclass()
class AciData:
    tenant: str
    action: str
    protocol: str
    src_port: str
    src_portgroup: str
    dst_port: str
    dst_portgroup: str
    src_epg: str
    dst_epg: str
    src_ap: str
    dst_ap: str
    subject_description: str

    def __post_init__(self):
        self.contract_name = f"{self.src_epg}_to_{self.dst_epg}_Con"

    def filters(self):

        my_filter = AciFilter(
            protocol=self.protocol,
            src_port=self.src_port,
            dst_port=self.dst_port,
            src_portgroup=self.src_portgroup,
            dst_portgroup=self.dst_portgroup,
        )

        return my_filter


@dataclass()
class AciSubject:
    filter_name: str
    action: str
    protocol: str
    description: str
    name: str = None

    def __post_init__(self):
        # Set the Subject name if not already set
        if not self.name:
            self.name = f"{self.action.capitalize()}_{self.filter_name}"
            self.name = self.name.replace("_Fltr", "_Subj")

    def json(self):
        if self.protocol == "UDP":
            j2_subject = j2_env.get_template("new_subject_udp.jinja2")
        else:
            j2_subject = j2_env.get_template("new_subject.jinja2")

        kwargs = {
            "SUBJECT_NAME": self.name,
            "SUBJECT_DESCRIPTION": self.description,
            "FILTER_NAME": self.filter_name,
            "ACTION": self.action,
        }
        new_subject = j2_subject.render(**kwargs)
        return new_subject


@dataclass()
class AciFilterRule:
    name: str
    protocol: str
    src_port_from: str
    src_port_to: str
    dst_port_from: str
    dst_port_to: str

    def json(self):
        j2_rule = j2_env.get_template("new_filter_rule.jinja2")

        kwargs = {
            "RULE_NAME": self.name,
            "PROTOCOL": self.protocol,
            "SRC_PORT_FROM": self.src_port_from,
            "SRC_PORT_TO": self.src_port_to,
            "DST_PORT_FROM": self.dst_port_from,
            "DST_PORT_TO": self.dst_port_to,
        }

        new_rule = j2_rule.render(**kwargs)

        return new_rule


@dataclass()
class AciFilter:
    protocol: str
    src_port: str
    dst_port: str
    name: str = None
    src_portgroup: str = None
    dst_portgroup: str = None
    rules: list[AciFilterRule] = field(default_factory=list)

    def __post_init__(self):
        # Set the filter name if not already set
        if not self.name:
            self.name = f"{self.protocol}_"
            if self.src_portgroup:
                self.name += f"Src_{self.src_portgroup}_"
            elif self.src_port:
                self.name += f"Src_{self.src_port}_"
            if self.dst_portgroup:
                self.name += f"Dst_{self.dst_portgroup}_"
            elif self.dst_port:
                self.name += f"Dst_{self.dst_port}_"
            self.name += "Fltr"
            self.name = self.name.replace(" ", "")

        self.set_rules()

    # this is SUPER ugly!!! fix?!??!
    def set_rules(self):
        # Default values
        base_rule_name = self.name.replace("_Fltr", "")
        src_port_from = "unspecified"
        src_port_to = "unspecified"
        dst_port_from = "unspecified"
        dst_port_to = "unspecified"

        if self.protocol not in ("TCP", "UDP"):
            # Supported ACI IPv4 Protocols:
            if self.protocol in (
                "EIGRP",
                "EGP",
                "ICMP",
                "IGMP",
                "IGP",
                "L2TP",
                "OSFPIGP",
                "PIM",
            ):
                protocol = self.protocol.lower()
            elif self.protocol == "IP":
                protocol = "unspecified"
            else:
                print(
                    f"Protocol {self.protocol} is not supported in ACI, please fix contract spreadsheet."
                )
                sys.exit()
            # Add a rule to a non tcp/udp filter
            self.add_rule(
                base_rule_name,
                protocol,
                src_port_from,
                src_port_to,
                dst_port_from,
                dst_port_to,
            )
        else:
            # Add a rule to a 'blank/no ports' filter
            if not self.src_port and not self.dst_port:
                self.add_rule(
                    base_rule_name,
                    self.protocol.lower(),
                    src_port_from,
                    src_port_to,
                    dst_port_from,
                    dst_port_to,
                )
            protocol = self.protocol.lower()

        # SOURCE
        if self.src_port:
            if "," in self.src_port:
                ports = self.src_port.split(",")
                for port in ports:
                    extra_rule_name = f"{base_rule_name}_{port}"
                    if "-" in port:
                        src_port_from, src_port_to = port.split(" - ")
                    else:
                        src_port_from = src_port_to = port
                    if self.dst_port:
                        if "," in self.dst_port:
                            extra_extra_rule_name = f"{extra_rule_name}_to"
                            dst_ports = self.dst_port.split(",")
                            for dst_port in dst_ports:
                                extra_extra_rule_name += f"_{dst_port}"
                                if "-" in dst_port:
                                    dst_port_from, dst_port_to = dst_port.split(" - ")
                                else:
                                    dst_port_from = dst_port_to = dst_port
                                self.add_rule(
                                    extra_extra_rule_name.replace(" ", ""),
                                    protocol,
                                    src_port_from,
                                    src_port_to,
                                    dst_port_from,
                                    dst_port_to,
                                )
                        else:
                            extra_rule_name += f"_to_{self.dst_port}"
                            if "-" in self.dst_port:
                                dst_port_from, dst_port_to = self.dst_port.split(" - ")
                            else:
                                dst_port_from = dst_port_to = self.dst_port
                    self.add_rule(
                        extra_rule_name.replace(" ", ""),
                        protocol,
                        src_port_from,
                        src_port_to,
                        dst_port_from,
                        dst_port_to,
                    )
            else:
                rule_name = f"{base_rule_name}"
                if "-" in self.src_port:
                    src_port_from, src_port_to = self.src_port.split(" - ")
                else:
                    src_port_from = src_port_to = self.src_port

                if self.dst_port:
                    if "," in self.dst_port:
                        dst_ports = self.dst_port.split(",")
                        for dst_port in dst_ports:
                            if "-" in dst_port:
                                dst_port_from, dst_port_to = dst_port.split(" - ")
                            else:
                                dst_port_from = dst_port_to = dst_port
                    else:
                        if "-" in self.dst_port:
                            dst_port_from, dst_port_to = self.dst_port.split(" - ")
                        else:
                            dst_port_from = dst_port_to = self.dst_port
                self.add_rule(
                    rule_name.replace(" ", ""),
                    protocol,
                    src_port_from,
                    src_port_to,
                    dst_port_from,
                    dst_port_to,
                )

        # DEST
        elif self.dst_port:
            if "," in self.dst_port:
                ports = self.dst_port.split(",")
                for port in ports:
                    extra_rule_name = f"{base_rule_name}_{port}"
                    if "-" in port:
                        dst_port_from, dst_port_to = port.split(" - ")
                    else:
                        dst_port_from = dst_port_to = port
                    self.add_rule(
                        extra_rule_name.replace(" ", ""),
                        protocol,
                        src_port_from,
                        src_port_to,
                        dst_port_from,
                        dst_port_to,
                    )
            else:
                rule_name = f"{base_rule_name}"
                if "-" in self.dst_port:
                    dst_port_from, dst_port_to = self.dst_port.split(" - ")
                else:
                    dst_port_from = dst_port_to = self.dst_port
                self.add_rule(
                    rule_name.replace(" ", ""),
                    protocol,
                    src_port_from,
                    src_port_to,
                    dst_port_from,
                    dst_port_to,
                )

    def add_rule(
        self,
        rule_name,
        protocol,
        src_port_from,
        src_port_to,
        dst_port_from,
        dst_port_to,
    ):
        rule = AciFilterRule(
            **{
                "name": f"{rule_name}",
                "protocol": protocol,
                "src_port_from": src_port_from,
                "src_port_to": src_port_to,
                "dst_port_from": dst_port_from,
                "dst_port_to": dst_port_to,
            }
        )
        self.rules.append(rule)

    def json(self):
        # return filter in json
        j2_filter = j2_env.get_template("base_filter.jinja2")
        rules = [rule.json() for rule in self.rules]
        new_filter = j2_filter.render(FILTER_NAME=self.name, ITEMS=rules)

        return new_filter


@dataclass()
class AciContract:
    name: str
    filter_name: str = None
    subjects: List[AciSubject] = field(default_factory=list)

    def __contains__(self, subject):
        for my_subject in self.subjects:
            if my_subject.name == subject.name:
                return True

    def json(self):
        j2_contract = j2_env.get_template("new_contract.jinja2")
        subjects_output = [subject.json() for subject in self.subjects]
        new_contract = j2_contract.render(
            CONTRACT_NAME=self.name, SUBJECTS=subjects_output
        )
        return new_contract


@dataclass()
class AciTenant:
    name: str
    filters: List[AciFilter] = field(default_factory=list)
    contracts: List[AciContract] = field(default_factory=list)

    def __post_init__(self):
        if self.filters is None or self.contracts is None:
            print("Error, Unable to initialize tenant")
            sys.exit()

    def json(self):
        pass

    def to_file(self, filename: str) -> None:
        j2_tenant_base = j2_env.get_template("base_tenant.jinja2")
        filter_output = [a_filter.json() for a_filter in self.filters]
        contract_ouput = [contract.json() for contract in self.contracts]
        children = filter_output + contract_ouput
        with open(filename, "w") as fout:
            fout.write(j2_tenant_base.render(CHILDREN=children))

        print(f"{filename} created.")

    #
    #
    # items = [filter.json() for filter in new_filters]
    # for contract_name in new_subjects.keys():
    #     subjects = new_subjects[contract_name]["subjects"]
    #     items.append(j2_contract_base.render(CONTRACT_NAME=contract_name, SUBJECTS=subjects))


@dataclass()
class CustomWorksheet:
    def __init__(self, excel_filename):
        # Load excel and get first sheet/tab
        _ = load_excel(excel_filename)
        self.worksheet: Workbook = _.active

        # Get headers and put into dict so that sheet["headername"] = "column_letter"
        self.col_dict = {}
        for col in self.worksheet.iter_cols(1, self.worksheet.max_column):
            self.col_dict[col[0].value.upper()] = col[0].column_letter

    def get_contracts(self):
        contracts = []
        for row, _ in enumerate(self.worksheet.iter_rows(), start=1):
            if row == 1:
                continue
            subject_description = self.worksheet[
                self.col_dict["REMARK"] + str(row)
            ].value
            if not subject_description:
                subject_description = ""

            src_port = self.worksheet[self.col_dict["SRC_PORT"] + str(row)].value
            dst_port = self.worksheet[self.col_dict["DST_PORT"] + str(row)].value
            src_port = str(src_port) if src_port else None
            dst_port = str(dst_port) if dst_port else None

            data = {
                "tenant": self.worksheet[self.col_dict["SRC_TENANT"] + str(row)].value,
                "action": self.worksheet[self.col_dict["ACTION"] + str(row)].value,
                "protocol": self.worksheet[
                    self.col_dict["PROTOCOL"] + str(row)
                ].value.upper(),
                "src_port": src_port,
                "src_portgroup": self.worksheet[
                    self.col_dict["SRC_PORTGROUP"] + str(row)
                ].value,
                "dst_port": dst_port,
                "dst_portgroup": self.worksheet[
                    self.col_dict["DST_PORTGROUP"] + str(row)
                ].value,
                "src_epg": self.worksheet[self.col_dict["SRC_EPG"] + str(row)].value,
                "src_ap": self.worksheet[
                    self.col_dict["SRC_APPLICATION"] + str(row)
                ].value,
                "dst_epg": self.worksheet[self.col_dict["DST_EPG"] + str(row)].value,
                "dst_ap": self.worksheet[
                    self.col_dict["DST_APPLICATION"] + str(row)
                ].value,
                "subject_description": subject_description,
            }
            aci_data = AciData(**data)

            contracts.append(aci_data)

        return contracts
